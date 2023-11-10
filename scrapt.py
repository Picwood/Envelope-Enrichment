from selenium import webdriver
import openpyxl
import re
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Set up the Chrome driver options
options = Options()
options.add_argument("--headless")  # Run Chrome in headless mode
options.add_argument("--disable-gpu")
driver = webdriver.Chrome(options=options)  # Adjust path

# Load the Excel file
wb = openpyxl.load_workbook(r'C:\Users\flori\Downloads\COTERIE 2.xlsx')
sheet = wb.active

# Regex for email detection
email_pattern = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

for row in range(2, 660): #sheet.max_row + 1
    try:
        url = sheet.cell(row=row, column=2).value
        driver.get(url)
    except:
        continue    

    # Try to find the link using the class name you mentioned
    try:
    # Wait up to 10 seconds until the link element is located
        site_contact_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "contact")))
        print(site_contact_element)
    except:
        email_contact = "N\A"
    else:
        email_contact = site_contact_element.get_attribute('href')
        print(email_contact)

        sheet.cell(row=row, column=3).value = email_contact

# Save the changes
wb.save(r'C:\Users\flori\Downloads\COTERIE 3.xlsx')
driver.quit()

#MuiTypography-root MuiTypography-subtitle2 MuiTypography-noWrap css-101vmsx
